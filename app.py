import streamlit as st
import pandas as pd
import psycopg2
from datetime import datetime, timezone, timedelta
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# Importar para formatação do Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(
    page_title="Salim Outlet - Controle de Motos", 
    layout="wide",
    initial_sidebar_state="expanded"  # Sidebar sempre visível
)

# Fuso horário de Brasília
fuso_brasilia = timezone(timedelta(hours=-3))

# Inicializar sessão
if 'chassis' not in st.session_state:
    st.session_state.chassis = []
if 'last_chassi' not in st.session_state:
    st.session_state.last_chassi = ""
if 'input_key' not in st.session_state:
    st.session_state.input_key = 0

def conectar_banco():
    """Conecta ao banco Neon"""
    try:
        conn = psycopg2.connect(
            host=st.secrets["NEON_HOST"],
            database=st.secrets["NEON_DATABASE"],
            user=st.secrets["NEON_USER"],
            password=st.secrets["NEON_PASSWORD"],
            port=st.secrets["NEON_PORT"],
            sslmode='require'
        )
        return conn
    except Exception as e:
        st.error(f"Erro de conexão: {str(e)}")
        return None

def criar_excel_formatado(df, operador):
    """Cria um Excel formatado com duas abas: Listagem Completa e Sumário por SKU"""
    
    # Criar workbook
    wb = Workbook()
    
    # Remover aba padrão criada automaticamente
    wb.remove(wb.active)
    
    # ===== ABA 1: LISTAGEM COMPLETA =====
    ws_lista = wb.create_sheet("Listagem Completa")
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    success_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Verde claro
    error_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")    # Vermelho claro
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Título principal
    ws_lista.merge_cells('A1:F1')
    ws_lista['A1'] = f"RELATÓRIO DE CONTAGEM DE CHASSI - SALIM OUTLET"
    ws_lista['A1'].font = Font(bold=True, size=16, color="2E86AB")
    ws_lista['A1'].alignment = center_align
    
    # Informações da loja e data
    ws_lista.merge_cells('A2:F2')
    ws_lista['A2'] = f"Loja: {operador} - Data: {datetime.now(fuso_brasilia).strftime('%d/%m/%Y %H:%M')}"
    ws_lista['A2'].font = Font(bold=True, size=12)
    ws_lista['A2'].alignment = center_align
    
    # Estatísticas
    total = len(df)
    encontrados = len(df[df['status'] == 'Encontrado'])
    nao_encontrados = len(df[df['status'] == 'Não encontrado'])
    
    ws_lista.merge_cells('A3:F3')
    ws_lista['A3'] = f"RESUMO: Total: {total} | Encontrados: {encontrados} | Não Encontrados: {nao_encontrados}"
    ws_lista['A3'].font = Font(bold=True, size=11, color="FFD700")  # Amarelo ouro
    ws_lista['A3'].alignment = center_align
    ws_lista['A3'].fill = PatternFill(start_color="2E2E2E", end_color="2E2E2E", fill_type="solid")
    
    # Espaço em branco
    ws_lista['A4'] = ""
    
    # Cabeçalhos das colunas
    headers = ['Chassi', 'Data', 'Descrição', 'Modelo (SKU)', 'Montador', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws_lista.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
        # Ajustar largura das colunas
        ws_lista.column_dimensions[chr(64 + col)].width = 20
    
    # Dados
    for row_idx, (index, row) in enumerate(df.iterrows(), 6):
        for col_idx, value in enumerate([row['chassi'], row['data'], row['descricao'], 
                                       row['modelo'], row['montador'], row['status']], 1):
            cell = ws_lista.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Colorir baseado no status
            if row['status'] == 'Encontrado':
                cell.fill = success_fill
            else:
                cell.fill = error_fill
    
    # Adicionar totais na última linha
    last_row = len(df) + 7
    ws_lista.merge_cells(f'A{last_row}:E{last_row}')
    ws_lista[f'A{last_row}'] = "TOTAIS:"
    ws_lista[f'A{last_row}'].font = Font(bold=True)
    ws_lista[f'A{last_row}'].alignment = Alignment(horizontal='right')
    
    ws_lista[f'F{last_row}'] = f"Total: {total}"
    ws_lista[f'F{last_row}'].font = Font(bold=True)
    ws_lista[f'F{last_row}'].border = border
    ws_lista[f'F{last_row}'].alignment = center_align

    # ===== ABA 2: SUMÁRIO POR SKU =====
    ws_sumario = wb.create_sheet("Sumário por SKU")
    
    # Criar dataframe com sumário por SKU
    df_sumario = df[df['status'] == 'Encontrado'].groupby(['modelo', 'descricao']).size().reset_index()
    df_sumario.columns = ['SKU', 'Descrição', 'Quantidade']
    df_sumario = df_sumario.sort_values('Quantidade', ascending=False)
    
    # Título da aba de sumário
    ws_sumario.merge_cells('A1:C1')
    ws_sumario['A1'] = f"SUMÁRIO POR SKU - SALIM OUTLET"
    ws_sumario['A1'].font = Font(bold=True, size=16, color="2E86AB")
    ws_sumario['A1'].alignment = center_align
    
    # Informações da loja e data
    ws_sumario.merge_cells('A2:C2')
    ws_sumario['A2'] = f"Loja: {operador} - Data: {datetime.now(fuso_brasilia).strftime('%d/%m/%Y %H:%M')}"
    ws_sumario['A2'].font = Font(bold=True, size=12)
    ws_sumario['A2'].alignment = center_align
    
    # Total de SKUs diferentes
    total_skus = len(df_sumario)
    total_quantidade = df_sumario['Quantidade'].sum()
    
    ws_sumario.merge_cells('A3:C3')
    ws_sumario['A3'] = f"RESUMO: {total_skus} SKUs diferentes | Total de unidades: {total_quantidade}"
    ws_sumario['A3'].font = Font(bold=True, size=11, color="FFD700")
    ws_sumario['A3'].alignment = center_align
    ws_sumario['A3'].fill = PatternFill(start_color="2E2E2E", end_color="2E2E2E", fill_type="solid")
    
    # Espaço em branco
    ws_sumario['A4'] = ""
    
    # Cabeçalhos do sumário
    sumario_headers = ['SKU', 'Descrição', 'Quantidade']
    for col, header in enumerate(sumario_headers, 1):
        cell = ws_sumario.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
        # Ajustar largura das colunas
        ws_sumario.column_dimensions[chr(64 + col)].width = 25
    
    # Dados do sumário
    for row_idx, (index, row) in enumerate(df_sumario.iterrows(), 6):
        for col_idx, value in enumerate([row['SKU'], row['Descrição'], row['Quantidade']], 1):
            cell = ws_sumario.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Colorir baseado na quantidade (gradiente de verde)
            if row['Quantidade'] >= 10:
                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            elif row['Quantidade'] >= 5:
                cell.fill = PatternFill(start_color="8BC34A", end_color="8BC34A", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    
    # Adicionar totais na última linha do sumário
    last_row_sumario = len(df_sumario) + 7
    ws_sumario.merge_cells(f'A{last_row_sumario}:B{last_row_sumario}')
    ws_sumario[f'A{last_row_sumario}'] = "TOTAL GERAL:"
    ws_sumario[f'A{last_row_sumario}'].font = Font(bold=True)
    ws_sumario[f'A{last_row_sumario}'].alignment = Alignment(horizontal='right')
    
    ws_sumario[f'C{last_row_sumario}'] = total_quantidade
    ws_sumario[f'C{last_row_sumario}'].font = Font(bold=True, size=12)
    ws_sumario[f'C{last_row_sumario}'].border = border
    ws_sumario[f'C{last_row_sumario}'].alignment = center_align
    ws_sumario[f'C{last_row_sumario}'].fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    
    # Nome do arquivo
    filename = f"contagem_salim_outlet_{operador}_{datetime.now(fuso_brasilia).strftime('%Y%m%d_%H%M')}.xlsx"
    
    # Salvar arquivo
    wb.save(filename)
    return filename

def main():
    # Cabeçalho com cor AMARELO OURO, CONTORNO PRETO e TAMANHO DOBRADO
    st.markdown(
        """
        <h1 style='
            color: #FFD700; 
            margin-bottom: 20px; 
            font-size: 5rem;
            text-shadow: 
                -2px -2px 0 #000,
                2px -2px 0 #000,
                -2px 2px 0 #000,
                2px 2px 0 #000,
                -3px 0px 0 #000,
                3px 0px 0 #000,
                0px -3px 0 #000,
                0px 3px 0 #000;
            font-weight: bold;
            text-align: center;
        '>Salim Outlet - Controle de Scooters</h1>
        """,
        unsafe_allow_html=True
    )
    
    st.divider()
    
    # Área principal - Formulário de chassis
    st.header("📝 Registrar Chassi")
    
    # Container para o campo de chassi
    chassi_container = st.container()
    
    with chassi_container:
        # Campo de chassi com key dinâmica
        chassi = st.text_input(
            "Digite o número do chassi ou use leitor de código de barras:",
            placeholder="⬅️ POSICIONE O LEITOR AQUI - O CAMPO ESTÁ PRONTO",
            key=f"chassi_input_{st.session_state.input_key}",
            label_visibility="visible"
        )
    
    # JavaScript MUITO SIMPLES - apenas tenta focar uma vez
    st.markdown("""
    <script>
        // Espera a página carregar e tenta focar no campo
        setTimeout(function() {
            // Procura por inputs com placeholder que contenha "leitor"
            const inputs = document.querySelectorAll('input');
            for (let input of inputs) {
                if (input.placeholder && input.placeholder.includes('LEITOR')) {
                    input.focus();
                    input.select();
                    console.log('Campo de chassi focado');
                    break;
                }
            }
        }, 1000);
    </script>
    """, unsafe_allow_html=True)
    
    # Verifica se há um novo chassi para registrar (modo automático)
    if (chassi and 
        chassi.strip() and 
        chassi != st.session_state.last_chassi):
        
        st.session_state.last_chassi = chassi
        registrar_chassi(chassi.strip())
        # Incrementa a key para forçar novo campo limpo
        st.session_state.input_key += 1
        # Força o rerun para limpar o campo
        st.rerun()

    # Instruções para uso com leitor de código de barras
    st.success("""
    **🎯 MODO LEITOR DE CÓDIGO DE BARRAS ATIVADO**
    
    **→ POSICIONE O LEITOR NO CAMPO ACIMA ←**
    
    - ✅ **Gravação automática** a cada leitura  
    - ✅ **Campo limpo** após cada registro
    - ✅ **Pronto para próxima leitura**
    
    *Dica: Se o campo não estiver com foco, clique uma vez nele e depois use o leitor.*
    """)

    # Sidebar FIXA
    with st.sidebar:
        # Logo na sidebar (MANTIDO)
        st.image("salimoutlet.jpg", width=100)
        
        st.divider()
        
        # Campo para nome da loja
        operador = st.text_input(
            "🏪 Loja:",
            placeholder="Digite o nome da loja",
            key="operador_input"
        )
        
        # Contador
        st.metric("📋 Chassis Registrados", len(st.session_state.chassis))
        
        st.divider()
        
        # Informação do modo automático
        st.info("🔴 **Modo Leitor Ativo**")
        st.caption("Gravação automática ao ler código de barras")
        
        # Botão de nova contagem
        if st.button("🔄 Nova Contagem", use_container_width=True, type="secondary"):
            st.session_state.chassis = []
            st.session_state.last_chassi = ""
            st.session_state.input_key += 1
            st.rerun()
        
        st.divider()
        
        # Botão finalizar (só aparece se tiver chassis)
        if st.session_state.chassis:
            if st.button("✅ FINALIZAR CONTAGEM", use_container_width=True, type="primary"):
                if operador:
                    finalizar_automático(operador)
                else:
                    st.warning("⚠️ Digite o nome da loja")

    # Lista de chassis registrados
    if st.session_state.chassis:
        st.header("📋 Chassis Registrados")
        
        # DataFrame com formatação
        df = pd.DataFrame(st.session_state.chassis)
        st.dataframe(df, use_container_width=True, hide_index=True)
        
        # Estatísticas rápidas
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total", len(st.session_state.chassis))
        with col2:
            encontrados = len([c for c in st.session_state.chassis if c['status'] == 'Encontrado'])
            st.metric("Encontrados", encontrados)
        with col3:
            nao_encontrados = len([c for c in st.session_state.chassis if c['status'] == 'Não encontrado'])
            st.metric("Não Encontrados", nao_encontrados)
            
        # Aviso sobre finalização
        if not st.session_state.get('operador_input'):
            st.warning("👆 **Digite o nome da loja na sidebar para finalizar**")

def registrar_chassi(chassi_numero):
    """Registra um chassi"""
    if not chassi_numero:
        return
        
    # Verificar duplicado
    if any(c['chassi'] == chassi_numero for c in st.session_state.chassis):
        st.warning(f"⚠️ Chassi {chassi_numero} já foi registrado!")
        return
    
    # Consultar banco
    conn = conectar_banco()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("SELECT descricao, sku, montador FROM producao WHERE chassi = %s", (chassi_numero,))
            resultado = cur.fetchone()
            
            if resultado:
                descricao, modelo, montador = resultado
                registro = {
                    'chassi': chassi_numero,
                    'data': datetime.now(fuso_brasilia).strftime("%d/%m/%Y %H:%M"),
                    'descricao': descricao,
                    'modelo': modelo,
                    'montador': montador,
                    'status': 'Encontrado'
                }
                st.success(f"✅ **{chassi_numero}** - {descricao}")
            else:
                registro = {
                    'chassi': chassi_numero,
                    'data': datetime.now(fuso_brasilia).strftime("%d/%m/%Y %H:%M"),
                    'descricao': 'Não encontrado',
                    'modelo': 'N/A',
                    'montador': 'N/A',
                    'status': 'Não encontrado'
                }
                st.error(f"❌ **{chassi_numero}** - Não encontrado")
            
            st.session_state.chassis.append(registro)
            cur.close()
            conn.close()
            
        except Exception as e:
            st.error(f"Erro na consulta: {str(e)}")
    else:
        st.error("❌ Erro de conexão com o banco")

def finalizar_automático(operador):
    """Finaliza automaticamente - gera Excel e envia email"""
    try:
        # Gerar Excel FORMATADO com duas abas
        df = pd.DataFrame(st.session_state.chassis)
        filename = criar_excel_formatado(df, operador)
        
        # Enviar email automático
        enviar_email_automatico(filename, operador, df)
        
        # Mostrar sucesso
        st.balloons()
        st.success("🎉 **CONTAGEM FINALIZADA COM SUCESSO!**")
        
        # Estatísticas finais
        encontrados = len([c for c in st.session_state.chassis if c['status'] == 'Encontrado'])
        nao_encontrados = len([c for c in st.session_state.chassis if c['status'] == 'Não encontrado'])
        
        # Calcular estatísticas do sumário por SKU
        df_sumario = df[df['status'] == 'Encontrado'].groupby(['modelo', 'descricao']).size().reset_index()
        total_skus = len(df_sumario)
        total_unidades = df_sumario[0].sum() if len(df_sumario) > 0 else 0
        
        st.info(f"""
        **📊 Relatório enviado:**
        - **📧 Email:** Enviado automaticamente
        - **🏪 Loja:** {operador}
        - **📦 Total de chassis:** {len(st.session_state.chassis)}
        - **✅ Encontrados:** {encontrados}
        - **❌ Não encontrados:** {nao_encontrados}
        - **📈 SKUs diferentes:** {total_skus}
        - **📄 Excel:** 2 abas (Listagem Completa + Sumário por SKU)
        """)
        
        # Botão para baixar Excel
        with open(filename, "rb") as f:
            st.download_button(
                "📥 BAIXAR PLANILHA EXCEL COMPLETA",
                f,
                filename,
                "application/vnd.ms-excel",
                use_container_width=True,
                type="primary"
            )
            
    except Exception as e:
        st.error(f"❌ Erro ao finalizar: {str(e)}")

def enviar_email_automatico(arquivo, operador, df):
    """Envia email automaticamente"""
    try:
        # Verificar se as configurações de email existem
        required_secrets = ["EMAIL_FROM", "EMAIL_PASSWORD", "SMTP_SERVER", "SMTP_PORT"]
        missing_secrets = [secret for secret in required_secrets if secret not in st.secrets]
        
        if missing_secrets:
            st.warning(f"⚠️ Email não configurado. Faltando: {', '.join(missing_secrets)}")
            return False
        
        # Lista de emails fixa
        emails_destino = st.secrets.get("EMAIL_TO", "contagem.salimoutlet@gmail.com").split(",")
        emails_destino = [email.strip() for email in emails_destino if email.strip()]
        
        # Preparar email - ASSUNTO DINÂMICO COM NOME DA LOJA
        msg = MIMEMultipart()
        msg['From'] = st.secrets["EMAIL_FROM"]
        msg['To'] = ", ".join(emails_destino)
        msg['Subject'] = f"Relatório de Contagem - {operador} - {datetime.now(fuso_brasilia).strftime('%d/%m/%Y')}"
        
        # Estatísticas para o email
        encontrados = len([c for c in st.session_state.chassis if c['status'] == 'Encontrado'])
        nao_encontrados = len([c for c in st.session_state.chassis if c['status'] == 'Não encontrado'])
        df_sumario = df[df['status'] == 'Encontrado'].groupby(['modelo', 'descricao']).size().reset_index()
        total_skus = len(df_sumario)
        total_unidades = df_sumario[0].sum() if len(df_sumario) > 0 else 0
        
        # Corpo do email
        body = f"""
        RELATÓRIO DE CONTAGEM DE CHASSI - SALIM OUTLET
        
        Data: {datetime.now(fuso_brasilia).strftime('%d/%m/%Y %H:%M')}
        Loja: {operador}
        
        RESUMO:
        • Total de chassis: {len(df)}
        • Encontrados: {encontrados}
        • Não encontrados: {nao_encontrados}
        • SKUs diferentes: {total_skus}
        • Total de unidades: {total_unidades}
        
        O arquivo Excel em anexo contém:
        - ABA 1: Listagem completa de todos os chassis
        - ABA 2: Sumário por SKU com quantidades
        
        --
        Sistema de Controle de Chassi
        Salim Outlet
        """
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Anexar arquivo
        with open(arquivo, "rb") as f:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{arquivo}"')
        msg.attach(part)
        
        # Enviar email
        try:
            server = smtplib.SMTP_SSL(st.secrets["SMTP_SERVER"], int(st.secrets["SMTP_PORT"]))
            server.login(st.secrets["EMAIL_FROM"], st.secrets["EMAIL_PASSWORD"])
            server.send_message(msg)
            server.quit()
        except:
            server = smtplib.SMTP(st.secrets["SMTP_SERVER"], int(st.secrets["SMTP_PORT"]))
            server.starttls()
            server.login(st.secrets["EMAIL_FROM"], st.secrets["EMAIL_PASSWORD"])
            server.send_message(msg)
            server.quit()
        
        return True
        
    except Exception as e:
        st.error(f"❌ Erro no envio de email: {str(e)}")
        return False

if __name__ == "__main__":
    main()